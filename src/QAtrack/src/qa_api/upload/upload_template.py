from abc import ABC, abstractmethod
from collections import defaultdict
import os
import pandas as pd
from openpyxl import load_workbook


class QATrackUpload(ABC):
    def __init__(self, file_name: str, file_template: str):
        """

        Initialize QATrackUpload class

        Args:
            file_name (str): Name of the file
            file_root (str): Source root/directory of files (can find it if in deeper directory)
            file_template (str): Template of the file (for data grabbing)
        """
        self.file_name = file_name
        self.file_template = file_template

    @abstractmethod
    def setup_machine_template(self):
        """
        Setup machine template for QATrack+ API
        """
        pass

    @abstractmethod
    def setup_machine(self):
        """
        Setup machine for QATrack+ API
        """
        pass

    @abstractmethod
    def setup_template_sheet_data(self):
        pass

    def read_excel_sheets(
        self, type: str, template_file_name: str, machine_file_name: str
    ) -> pd.DataFrame:
        """
        Reads the excel sheet and returns a dataframe of main.

        Args:
            file_name (str): _description_

        Returns:
            pd.DataFrame: _description_
        """
        # Load your workbook
        template_workbook = load_workbook(filename=template_file_name)
        machine_workbook = load_workbook(filename=machine_file_name)
        # Get your sheets
        sheet1 = template_workbook[f"QATrack_{type}"]
        sheet2 = machine_workbook[type]

        # Initialize an empty dictionary to store names, corresponding values and cell locations
        data = defaultdict(dict)

        # Iterate through all cells in sheet1
        for row in sheet1.iter_rows():
            for cell in row:
                # Check if cell is not empty and contains a string
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and cell.value.startswith("#!")
                ):
                    # Get corresponding value from same cell in sheet2
                    value = sheet2[cell.coordinate].value
                    # Store name and value in the dictionary
                    data[cell.coordinate]["name"] = cell.value[2:]
                    data[cell.coordinate]["value"] = value
                    data[cell.coordinate]["location"] = cell.coordinate

        # Turn data into a dataframe
        df = pd.DataFrame(data.values())
        return df

    def find_file_recursive(self, directory, target_filename) -> str:
        """
        Find file recursively in directory

        Args:
            directory (_type_): Target directory
            target_filename (_type_): Target filename

        Returns:
            str: Path to file if found
        """
        for root, dirs, files in os.walk(directory):
            if target_filename in files:
                return os.path.join(root, target_filename)

        # Throw error if not found
        raise FileNotFoundError(f"File {target_filename} not found in {directory}")
